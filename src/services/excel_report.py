import os
import openpyxl
from datetime import datetime, timezone
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.db.models import User, SpreadHistory, Payment

REPORTS_DIR = "reports"
os.makedirs(REPORTS_DIR, exist_ok=True)

async def export_users(session: AsyncSession) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    
    headers = ["Telegram ID", "Статус", "Всего раскладов", "Средний стресс", "Доминирующая сфера", "Последний отчет", "Дата регистрации"]
    ws.append(headers)
    
    stmt = select(User).order_by(User.created_at.desc())
    result = await session.execute(stmt)
    users = result.scalars().all()
    
    for u in users:
        ws.append([
            u.telegram_id,
            u.subscription_status,
            u.total_spreads,
            round(u.avg_stress_index, 2) if u.avg_stress_index is not None else 0,
            "ИИ зашифровано", # Зашифровано
            u.last_report_date.strftime("%Y-%m-%d %H:%M") if u.last_report_date else "Нет",
            u.created_at.strftime("%Y-%m-%d %H:%M")
        ])
        
    filepath = os.path.join(REPORTS_DIR, f"users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(filepath)
    return filepath

async def export_spreads(session: AsyncSession) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Расклады"
    
    # Мы не можем расшифровать имена пользователей просто так, поэтому выводим User ID БД
    headers = ["User ID", "Тема", "Карты", "Индекс стресса", "Дата"]
    ws.append(headers)
    
    stmt = select(SpreadHistory).order_by(SpreadHistory.created_at.desc())
    result = await session.execute(stmt)
    history = result.scalars().all()
    
    for h in history:
        ws.append([
            h.user_id,
            h.topic,
            h.cards,
            h.stress_index,
            h.created_at.strftime("%Y-%m-%d %H:%M")
        ])
        
    filepath = os.path.join(REPORTS_DIR, f"spreads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(filepath)
    return filepath

async def export_subscriptions(session: AsyncSession) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Подписки"
    
    headers = ["Telegram ID", "Дата начала", "Дата окончания", "Статус", "Всего раскладов"]
    ws.append(headers)
    
    stmt = select(User).where(User.subscription_status == "pro").order_by(User.subscription_end_date.desc())
    result = await session.execute(stmt)
    pro_users = result.scalars().all()
    
    now = datetime.now(timezone.utc)
    for u in pro_users:
        # Для простоты считаем дату начала как 'Конец - месяц' если нет истории
        # В идеале это тягается из таблицы платежей
        end_dt = u.subscription_end_date
        status = "Активно"
        if end_dt and end_dt < now:
            status = "Просрочено"
            
        ws.append([
            u.telegram_id,
            "-", # Начало
            end_dt.strftime("%Y-%m-%d %H:%M") if end_dt else "Бессрочно",
            status,
            u.total_spreads
        ])
        
    filepath = os.path.join(REPORTS_DIR, f"subscriptions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(filepath)
    return filepath

async def export_payments(session: AsyncSession) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Платежи"
    
    headers = ["User ID", "Тип", "Сумма", "Статус", "Дата"]
    ws.append(headers)
    
    stmt = select(Payment).order_by(Payment.created_at.desc())
    result = await session.execute(stmt)
    payments = result.scalars().all()
    
    for p in payments:
        ws.append([
            p.user_id,
            p.payment_type,
            p.amount,
            p.status,
            p.created_at.strftime("%Y-%m-%d %H:%M")
        ])
        
    filepath = os.path.join(REPORTS_DIR, f"payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(filepath)
    return filepath
